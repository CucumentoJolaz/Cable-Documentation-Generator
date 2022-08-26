import datetime
import os
from typing import Dict, List, Tuple
import openpyxl as xl
import xlwings as xw
import shutil
import pywintypes
import pythoncom
from django.http import HttpResponse

from config.settings import BASE_DIR


def get_cable_parameters(post_form: Dict) -> List[Dict]:
    """
    Form list with dict of cable parameters from the received POST form
    :return: List[Dict]
    """
    cable_type_name_list = sorted([cable_type for cable_type in post_form
                                   if cable_type.startswith("type")])
    cable_scheme_name_list = sorted([cable_type for cable_type in post_form
                                     if cable_type.startswith("scheme")])
    cable_length_name_list = sorted([cable_type for cable_type in post_form
                                     if cable_type.startswith("length")])
    cable_parameter_list = []
    for i, _ in enumerate(cable_type_name_list):
        cable_parameter_list.append(
            {
                "type": post_form[cable_type_name_list[i]],
                "scheme": post_form[cable_scheme_name_list[i]],
                "length": post_form[cable_length_name_list[i]],
            }
        )

    return cable_parameter_list


def check_or_create_folder(folder_path: str) -> bool:
    """
    Checking if folder exists and create it if not.
    If folder exists - return True, if folder created - return False.
    :return:
    """
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
        return False
    else:
        return True


def copy_xlsm_to_buffer() -> str:
    """
    Copying macro file to buffer workfolder.
     Returns path to the xlsm file in buffer folder.
    """
    str_base_dir = str(BASE_DIR)
    macro_file_name = str_base_dir + "/generator/static/generator/xlsx/cost_calc.xlsm"

    folder_name = f"{datetime.datetime.now().strftime('%m.%d.%Y %H.%M.%S')}"
    macro_work_folder = str_base_dir + "/media/buffer/" + folder_name
    check_or_create_folder(str_base_dir + "/media/buffer")
    os.mkdir(macro_work_folder)
    shutil.copy2(macro_file_name, f"{macro_work_folder}/cost_calc.xlsm")
    return f"{macro_work_folder}/cost_calc.xlsm"

def get_folder_path(full_path) -> Tuple[str, str]:
    """ Get full path to file and return  folder path and filename"""
    return "/".join(full_path.split("/")[:-1]), full_path.split("/")[-1]

def run_macros_calculations(xlsm_path, cable_parameter_list) -> None:
    """
    Modify macros file with data and run it in buffer folder
    :return:
    """

    calc_wb = xl.load_workbook(xlsm_path, read_only=False, keep_vba=True)
    ws2 = calc_wb.active

    for i, _ in enumerate(cable_parameter_list):
        ws2.cell(row=i + 6, column=2, value=f"{cable_parameter_list[i]['type']} {cable_parameter_list[i]['scheme']}")
        ws2.cell(row=i + 6, column=3, value=float(cable_parameter_list[i]['length']))

    calc_wb.save(filename=xlsm_path)
    pythoncom.CoInitialize()
    with xw.App(add_book=False) as app:
        try:
            wb = app.books.open(xlsm_path)
            run_macro = wb.app.macro("Module1.Cblcalc")
            run_macro()
        except pywintypes.com_error:
            pass  # bruh. Macros executes anyway. But django server continue to work. And this error do not lay it down





def archive_and_delete(buffer_folder_path: str) -> Tuple[str, str]:
    """
    Create archive from folder and delete folder
    :return:
    """
    str_base_dir = str(BASE_DIR)
    check_or_create_folder(str_base_dir + "/media/buffer")

    shutil.make_archive(base_name=buffer_folder_path,
                        format='zip',
                        root_dir=buffer_folder_path)
    shutil.rmtree(buffer_folder_path)
    general_buffer_path, buffer_folder_name = get_folder_path(buffer_folder_path)
    archive_name = f"{buffer_folder_name}.zip"
    archive_path = f"{general_buffer_path}/{archive_name}"
    return archive_path


def run_documentation_calculation(cable_parameter_list: List[Dict]) -> str:
    """
    Generating EXCEL files from cable_parameter_list, and return them to user in ZIP archive
    """

    # add data to workbook with macro

    xlsm_path = copy_xlsm_to_buffer()
    run_macros_calculations(xlsm_path, cable_parameter_list)

    buffer_folder_path, filename = get_folder_path(xlsm_path)

    return buffer_folder_path


def serve_archived_documentation(archive_path: str) -> HttpResponse:
    """
    Serving HttpResponse object filled with appropriate ZIP archive with calculated documentation for cables.
    """
    response = HttpResponse(open(archive_path, 'rb'), content_type='application/zip')
    _, archive_name = get_folder_path(archive_path)
    response['Content-Disposition'] = f'attachment; filename={archive_name}'

    return response
